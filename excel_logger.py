import os
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook

import sensor

# Name of the output Excel file
EXCEL_FILE = "sensor_data.xlsx"

def init_excel():
    # If the file does not exist, create a workbook and write the headers
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sensor Readings"
        
        # Write column headers in the first row
        ws.append(["Timestamp", "Temperature (C)", "Pressure (hPa)"])
        
        wb.save(EXCEL_FILE)
        print(f"Created new Excel file: {EXCEL_FILE}")

def log_to_excel():
    try:
        # 1. Read data from sensor
        data = sensor.read()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        temp = data["temperature_c"]
        press = data["pressure_hpa"]

        # 2. Load the existing workbook and select active sheet
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # 3. Append the new data row
        ws.append([timestamp, temp, press])

        # 4. Save and close the workbook
        wb.save(EXCEL_FILE)
        wb.close()

        print(f"Logged to Excel -> Time: {timestamp} | Temp: {temp} C | Press: {press} hPa")

    except Exception as e:
        print(f"Error logging data to Excel: {e}")

if __name__ == "__main__":
    init_excel()
    print("Starting Excel logging. Press Ctrl+C to stop...")
    while True:
        log_to_excel()
        time.sleep(5)  # Log every 5 seconds (adjust as needed)
